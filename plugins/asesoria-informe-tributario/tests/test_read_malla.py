import importlib.util, os
from openpyxl import Workbook
def _load(name):
    spec = importlib.util.spec_from_file_location(name,
        os.path.join(os.path.dirname(__file__), "..", "scripts", name + ".py"))
    m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m); return m
rm = _load("read_malla")

def _fake(tmp_path):
    wb = Workbook(); ws = wb.active
    ws["B5"] = "Empresa Falsa"; ws["B6"] = "95% Otra Falsa"; ws["B7"] = "5% Juan Perez"
    p = tmp_path / "malla.xlsx"; wb.save(p); return str(p)

def _fake_multibloque(tmp_path):
    wb = Workbook(); ws = wb.active
    # Bloque 1: matriz + sus socios directos (suman ~100)
    ws["B3"] = "Matriz Falsa"
    ws["B4"] = "96% Socio Falso Uno"
    ws["B5"] = "4% Persona Falsa"
    # Bloque 2: OTRA entidad donde la matriz aparece como socia (dirección inversa)
    ws["B7"] = "Otra Entidad Falsa"
    ws["B8"] = "30% Matriz Falsa"
    p = tmp_path / "malla_multi.xlsx"; wb.save(p); return str(p)

def _fake_coma(tmp_path):
    wb = Workbook(); ws = wb.active
    ws["B3"] = "Matriz Coma"; ws["B4"] = "12,5% Fulano"
    p = tmp_path / "malla_coma.xlsx"; wb.save(p); return str(p)

def test_lee_participaciones(tmp_path):
    r = rm.read_malla_file(_fake(tmp_path))
    assert r["matriz"] == "Empresa Falsa"
    socios = {p["socio"]: p["pct"] for p in r["participaciones"]}
    assert socios["Otra Falsa"] == 95.0
    assert socios["Juan Perez"] == 5.0

def test_acota_al_primer_bloque(tmp_path):
    r = rm.read_malla_file(_fake_multibloque(tmp_path))
    assert r["matriz"] == "Matriz Falsa"
    socios = {p["socio"]: p["pct"] for p in r["participaciones"]}
    # Solo socios del primer bloque
    assert socios == {"Socio Falso Uno": 96.0, "Persona Falsa": 4.0}
    # El segundo bloque NO se incluye (la matriz como socia de otra entidad)
    assert "Matriz Falsa" not in socios
    assert not any(p["socio"] == "Matriz Falsa" for p in r["participaciones"])
    # Los pct del primer bloque suman ~100
    assert abs(sum(p["pct"] for p in r["participaciones"]) - 100.0) < 0.001

def test_coma_decimal(tmp_path):
    r = rm.read_malla_file(_fake_coma(tmp_path))
    pct = r["participaciones"][0]["pct"]
    assert pct == 12.5
    assert isinstance(pct, float)
