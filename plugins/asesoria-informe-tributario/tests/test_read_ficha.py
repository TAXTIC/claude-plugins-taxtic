import importlib.util, os
from datetime import datetime
from openpyxl import Workbook

def _load(name):
    spec = importlib.util.spec_from_file_location(name,
        os.path.join(os.path.dirname(__file__), "..", "scripts", name + ".py"))
    m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m); return m

rf = _load("read_ficha")

def _fake_xlsx(tmp_path):
    wb = Workbook(); ws = wb.active
    ws["B2"] = "Nombre o razón social:"; ws["C2"] = "EMPRESA FALSA SPA"
    ws["B3"] = "RUT:"; ws["C3"] = "11111111-1"
    ws["B9"] = "Regimen tributario"; ws["C9"] = "14A"
    p = tmp_path / "ficha.xlsx"; wb.save(p); return str(p)

def test_lee_ficha(tmp_path):
    r = rf.read_ficha_file(_fake_xlsx(tmp_path))
    assert r["razon_social"] == "EMPRESA FALSA SPA"
    assert r["regimen"] == "14A"

def _fake_xlsx_completo(tmp_path):
    """Espeja el layout real: fila header de fechas + fila de valores debajo,
    y sección de representantes (header + filas) seguida de 'Socios vigentes'
    (que NO debe colarse como representante). Datos 100% inventados."""
    wb = Workbook(); ws = wb.active
    ws["B2"] = "Nombre o razón social:"; ws["C2"] = "EMPRESA FALSA SPA"
    ws["B3"] = "RUT:"; ws["C3"] = "11111111-1"
    ws["B4"] = "Domicilio:"; ws["C4"] = "CALLE FICTICIA 123"
    ws["B7"] = "Fecha constitución"
    ws["C7"] = "Fecha Inicio de actividades"
    ws["D7"] = "Término de giro"
    ws["B8"] = datetime(2015, 3, 10)
    ws["C8"] = datetime(2015, 4, 1)
    ws["D8"] = "NO"
    ws["B9"] = "Regimen tributario"; ws["C9"] = "14A"
    ws["B11"] = "Representantes legales vigentes"
    ws["B12"] = "Nombre"; ws["C12"] = "Rut"; ws["D12"] = "A partir de"
    ws["B13"] = "JUAN PEREZ FICTICIO"; ws["C13"] = "22222222-2"; ws["D13"] = datetime(2015, 4, 1)
    ws["B16"] = "Socios vigentes"
    ws["B17"] = "Nombre"; ws["C17"] = "Rut"; ws["D17"] = "Capital enterado M$"
    ws["B18"] = "SOCIO FICTICIO UNO"; ws["C18"] = "33333333-3"
    p = tmp_path / "ficha_completa.xlsx"; wb.save(p); return str(p)

def test_lee_fechas_y_representantes(tmp_path):
    r = rf.read_ficha_file(_fake_xlsx_completo(tmp_path))
    assert r["razon_social"] == "EMPRESA FALSA SPA"
    assert r["regimen"] == "14A"
    assert r["fecha_constitucion"] == "2015-03-10"
    assert r["inicio_actividades"] == "2015-04-01"
    assert r["termino_giro"] == "NO"
    assert len(r["representantes"]) == 1
    assert r["representantes"][0]["nombre"] == "JUAN PEREZ FICTICIO"
    assert r["representantes"][0]["rut"] == "22222222-2"
    nombres = [x["nombre"] for x in r["representantes"]]
    assert "SOCIO FICTICIO UNO" not in nombres

def _fake_xlsx_rut_duplicado(tmp_path):
    """La fila real 'RUT:' | valor está arriba; más abajo hay una tabla cuyo
    header contiene otra celda 'Rut'. El parser NO debe dejar que ese header
    posterior sobrescriba el rut real del cliente. Datos inventados."""
    wb = Workbook(); ws = wb.active
    ws["B2"] = "Nombre o razón social:"; ws["C2"] = "EMPRESA FALSA SPA"
    ws["B3"] = "RUT:"; ws["C3"] = "11111111-1"
    ws["B9"] = "Regimen tributario"; ws["C9"] = "14A"
    # tabla posterior (socios/sociedades) con un header "Rut" que NO debe pisar
    ws["B20"] = "Nombre"; ws["C20"] = "Rut"; ws["D20"] = "A partir de"
    ws["B21"] = "SOCIO FICTICIO"; ws["C21"] = "99999999-9"
    p = tmp_path / "ficha_rut_dup.xlsx"; wb.save(p); return str(p)

def test_rut_toma_primera_coincidencia(tmp_path):
    r = rf.read_ficha_file(_fake_xlsx_rut_duplicado(tmp_path))
    assert r["rut"] == "11111111-1"

def _fake_xlsx_socios_pegados(tmp_path):
    """'Socios vigentes' viene INMEDIATAMENTE después de la fila del representante,
    SIN fila en blanco entre medio. Fuerza que el corte de representantes dependa
    del chequeo 'socios vigentes' (no del break por fila vacía). Datos inventados."""
    wb = Workbook(); ws = wb.active
    ws["B2"] = "Nombre o razón social:"; ws["C2"] = "EMPRESA FALSA SPA"
    ws["B3"] = "RUT:"; ws["C3"] = "11111111-1"
    ws["B11"] = "Representantes legales vigentes"
    ws["B12"] = "Nombre"; ws["C12"] = "Rut"; ws["D12"] = "A partir de"
    ws["B13"] = "REP FICTICIO UNO"; ws["C13"] = "22222222-2"; ws["D13"] = datetime(2016, 1, 5)
    # Sin fila en blanco: la sección de socios arranca justo debajo del representante.
    ws["B14"] = "Socios vigentes"
    ws["B15"] = "Nombre"; ws["C15"] = "Rut"; ws["D15"] = "Capital enterado M$"
    ws["B16"] = "SOCIO FICTICIO UNO"; ws["C16"] = "33333333-3"
    p = tmp_path / "ficha_socios_pegados.xlsx"; wb.save(p); return str(p)

def test_excluye_socios_sin_fila_en_blanco(tmp_path):
    r = rf.read_ficha_file(_fake_xlsx_socios_pegados(tmp_path))
    assert len(r["representantes"]) == 1
    assert r["representantes"][0]["nombre"] == "REP FICTICIO UNO"
    nombres = [x["nombre"] for x in r["representantes"]]
    assert "SOCIO FICTICIO UNO" not in nombres
    assert "Socios vigentes" not in nombres
