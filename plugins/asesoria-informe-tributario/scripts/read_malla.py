import re, sys, json
from openpyxl import load_workbook

def read_malla_file(path):
    """Lee una malla societaria (xlsx de texto libre) y devuelve solo los socios
    DIRECTOS de la matriz.

    El xlsx tiene VARIOS bloques de entidades: el primero es la matriz y sus
    socios (ej. "Empresa Matriz" <- 95% Socio Mayoritario + varias personas al
    1% = 100%). Los bloques siguientes son OTRAS entidades donde la matriz
    aparece como socia ("qué posee la matriz", dirección inversa).

    `participaciones` = socios directos de la matriz (SOLO el primer bloque). NO
    incluye las participaciones de la matriz en otras entidades. Se corta al
    primer texto no-% que sigue a la matriz, porque ese texto inicia el bloque
    de otra entidad.

    Retorna: {"matriz": str|None, "participaciones": [{"socio","rut","pct"}],
              "sin_parsear": [str]}  (sin_parsear = líneas no interpretadas,
              p.ej. líneas "%" que aparecen antes de la matriz).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    valores = [str(c).strip() for row in ws.iter_rows(values_only=True)
               for c in row if c is not None and str(c).strip()]
    matriz, participaciones, sin_parsear = None, [], []
    for v in valores:
        m = re.match(r"(\d+(?:[.,]\d+)?)\s*%\s+(.*)", v)
        if matriz is None:
            if m:
                # Línea "%" antes de identificar la matriz: anómalo. No la
                # descartamos en silencio, la registramos en sin_parsear.
                sin_parsear.append(v)
            else:
                matriz = v  # primer texto no-% = matriz (inicio del primer bloque)
            continue
        if m:
            participaciones.append({"socio": m.group(2).strip(), "rut": None,
                                     "pct": float(m.group(1).replace(",", "."))})
        else:
            # Texto no-% tras la matriz => inicia el bloque de OTRA entidad.
            # Cortamos: solo queremos los socios directos de la matriz.
            break
    return {"matriz": matriz, "participaciones": participaciones,
            "sin_parsear": sin_parsear}

if __name__ == "__main__":
    print(json.dumps(read_malla_file(sys.argv[1]), ensure_ascii=False, indent=2))
