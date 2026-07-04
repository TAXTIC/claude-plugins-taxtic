import sys, json
from datetime import date, datetime
from openpyxl import load_workbook

def _norm(s): return (str(s).strip().lower() if s is not None else "")

def _to_value(v):
    """Fechas -> ISO date string. Cualquier otro valor (ej. "NO") se deja tal cual."""
    if isinstance(v, datetime): return v.date().isoformat()
    if isinstance(v, date): return v.isoformat()
    return v

def read_ficha_file(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    out = {"razon_social": None, "rut": None, "domicilio": None, "regimen": None,
           "fecha_constitucion": None, "inicio_actividades": None, "termino_giro": None,
           "representantes": []}
    filas = list(ws.iter_rows(values_only=True))
    n = len(filas)

    i = 0
    while i < n:
        row = filas[i]

        # --- Pares label:valor adyacentes (razon_social, rut, domicilio, regimen) ---
        celdas = [c for c in row if c is not None]
        for j, c in enumerate(celdas[:-1]):
            k = _norm(c); val = celdas[j + 1]
            # Tomar la PRIMERA coincidencia (filas de arriba hacia abajo). No sobrescribir:
            # así la fila real (ej. B3 "RUT:" | valor) gana y los headers "Rut" de tablas
            # posteriores (representantes/socios/sociedades) no la pisan.
            if "raz" in k and "social" in k:
                if out["razon_social"] is None: out["razon_social"] = str(val).strip()
            elif k.startswith("rut"):
                if out["rut"] is None: out["rut"] = str(val).strip()
            elif "domicilio" in k:
                if out["domicilio"] is None: out["domicilio"] = str(val).strip()
            elif "regimen" in k:
                if out["regimen"] is None: out["regimen"] = str(val).strip()

        # --- Fila de headers de fechas (constitución / inicio actividades / término giro),
        #     cuyos valores están en la fila SIGUIENTE, posicionalmente por columna ---
        norm_by_col = {idx: _norm(v) for idx, v in enumerate(row) if v is not None}
        has_constituc = any("constituc" in v for v in norm_by_col.values())
        has_inicio = any("inicio de actividades" in v for v in norm_by_col.values())
        has_termino = any(("rmino" in v and "giro" in v) for v in norm_by_col.values())

        if has_constituc and has_inicio and has_termino and i + 1 < n:
            valores = filas[i + 1]
            for idx, v in norm_by_col.items():
                val = valores[idx] if idx < len(valores) else None
                if "constituc" in v: out["fecha_constitucion"] = _to_value(val)
                elif "inicio de actividades" in v: out["inicio_actividades"] = _to_value(val)
                elif "rmino" in v and "giro" in v: out["termino_giro"] = _to_value(val)
            i += 2
            continue

        # --- Sección "Representantes legales vigentes": header Nombre|Rut|A partir de
        #     y luego N filas de datos, hasta fila vacía o nueva sección (ej. Socios) ---
        if any("representantes legales vigentes" in v for v in norm_by_col.values()):
            j = i + 1
            while j < n and all(c is None for c in filas[j]):
                j += 1
            if j < n:
                header_row = filas[j]
                col_map = {}
                for idx, v in enumerate(header_row):
                    nv = _norm(v)
                    if nv == "nombre": col_map["nombre"] = idx
                    elif nv == "rut": col_map["rut"] = idx
                    elif "a partir de" in nv: col_map["desde"] = idx

                k = j + 1
                while k < n:
                    r = filas[k]
                    if all(c is None for c in r):
                        break
                    joined = " ".join(_norm(c) for c in r if c is not None)
                    if "socios vigentes" in joined:
                        break
                    nombre_idx = col_map.get("nombre")
                    nombre = r[nombre_idx] if nombre_idx is not None and nombre_idx < len(r) else None
                    if nombre is None:
                        break
                    rut_idx = col_map.get("rut")
                    rut = r[rut_idx] if rut_idx is not None and rut_idx < len(r) else None
                    desde_idx = col_map.get("desde")
                    desde = r[desde_idx] if desde_idx is not None and desde_idx < len(r) else None
                    out["representantes"].append({
                        "nombre": str(nombre).strip(),
                        "rut": str(rut).strip() if rut is not None else None,
                        "desde": _to_value(desde),
                    })
                    k += 1
                i = k
                continue

        i += 1

    return out

if __name__ == "__main__":
    print(json.dumps(read_ficha_file(sys.argv[1]), ensure_ascii=False, indent=2, default=str))
